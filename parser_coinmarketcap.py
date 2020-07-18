# -*- coding: utf-8 -*-
# https://coinmarketcap.com
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
import openpyxl


# pip install lxml
# pip install requests
# pip install beautifulsoup4
# pip install openpyxl


def get_html(url):
    while True:
        r = requests.get(url)   # Response
        if r.status_code != 200:
            print(f'r.status_code = {r.status_code}')
            time.sleep(30)
        else:
            break
    return r.text           # return html


def get_all_links(html):
    soup = BeautifulSoup(html, 'lxml')
    table = soup.find_all('div', class_='cmc-table__table-wrapper-outer')   # We get 3 div`s. Use last
    trs = table[-1].find_all('tr', class_='cmc-table-row')
    domain = 'https://coinmarketcap.com'
    links = []
    for tr in trs:
        rez = tr.find('div', 'cmc-table__column-name')
        rez = rez.find('a').get('href')     # string
        rez = domain + rez
        links.append(rez)
    return links


def get_page_data(html):
    soup = BeautifulSoup(html, 'lxml')
    try:
        div = soup.find('div', class_='cmc-details-panel-header')
    except:
        print('div not found')
        div = ''
    name = ''
    price = ''
    if div:
        try:
            name = div.find('h1').text.strip()
        except:
            print('name not found')
        try:
            price = div.find('span', class_='cmc-details-panel-price__price').text.strip()
        except:
            print('price not found')
    return name, price


def write_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    count = 1
    for i in data:
        ws.cell(row=count, column=1).value = i[0]
        ws.cell(row=count, column=2).value = i[1]
        count += 1
    ws.title = 'Result'
    wb.save(f'Parsing.xlsx')


def main():
    start = datetime.now()
    url = 'https://coinmarketcap.com'
    html = get_html(url)
    all_links = get_all_links(html)
    result = []
    for link in all_links:
        html = get_html(link)
        name, price = get_page_data(html)
        result.append((name, price))
        time.sleep(5)
        print(f'{name} {price}')
    write_xlsx(result)
    end = datetime.now()
    print(f'Total: {end - start}')


if __name__ == '__main__':
    main()
